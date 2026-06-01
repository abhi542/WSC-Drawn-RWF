from openpyxl import load_workbook

# ==========================================
# CONFIGURATION
# ==========================================

INPUT_FILE = "WSC-Drawn-RWF.xlsx"
OUTPUT_FILE = "updated_drawn_cumulative.xlsx"

# ==========================================
# LOAD WORKBOOK
# ==========================================

wb = load_workbook(INPUT_FILE)

pl_ws = wb["PL NO."]

# ==========================================
# BUILD SECTION -> COLUMN MAP
# ==========================================

header_row_1 = list(
    pl_ws.iter_rows(min_row=1, max_row=1, values_only=True)
)[0]

header_row_2 = list(
    pl_ws.iter_rows(min_row=2, max_row=2, values_only=True)
)[0]

col_map = {}

current_section = None

for col_idx, (h1, h2) in enumerate(zip(header_row_1, header_row_2)):

    if h1:
        current_section = str(h1).strip()

    if current_section and h2:
        col_map[(current_section, str(h2).strip().upper())] = col_idx

print(f"Found {len(col_map)} section/subheader mappings")

# ==========================================
# READ ALL MONTH SHEETS
# ==========================================

cumulative_data = {}

month_sheets = [
    sheet_name
    for sheet_name in wb.sheetnames
    if sheet_name != "PL NO."
]

print("\nMonthly sheets found:")
for sheet in month_sheets:
    print(f"  - {sheet}")

for sheet_name in month_sheets:

    ws = wb[sheet_name]

    headers = [
        str(x).strip() if x is not None else ""
        for x in next(
            ws.iter_rows(
                min_row=1,
                max_row=1,
                values_only=True
            )
        )
    ]

    try:
        section_idx = headers.index("SECTION")
        pl_idx = headers.index("PL NO.")
        drawn_idx = headers.index("DRAWN")
    except ValueError as e:
        print(f"Warning: Missing required column in sheet {sheet_name}. Error: {e}")
        continue

    row_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) <= max(section_idx, pl_idx, drawn_idx):
            continue

        section = row[section_idx]
        pl_no = row[pl_idx]
        drawn = row[drawn_idx]

        if not section or not pl_no:
            continue

        key = (
            str(pl_no).strip(),
            str(section).strip()
        )

        cumulative_data[key] = (
            cumulative_data.get(key, 0)
            + (drawn or 0)
        )

        row_count += 1

    print(f"Processed {row_count} rows from {sheet_name}")

# ==========================================
# UPDATE PL NO SHEET
# ==========================================

updated_rows = 0

for row in pl_ws.iter_rows(min_row=3):

    pl_no = row[1].value

    if not pl_no:
        continue

    pl_no = str(pl_no).strip()

    for (section, subheader), col_idx in col_map.items():

        if subheader != "DRAWN":
            continue

        key = (pl_no, section)

        if key in cumulative_data:

            row[col_idx].value = cumulative_data[key]
            updated_rows += 1

print(f"\nUpdated {updated_rows} DRAWN cells")

# ==========================================
# SAVE
# ==========================================

wb.save(OUTPUT_FILE)

print(f"\nDone!")
print(f"Output file: {OUTPUT_FILE}")